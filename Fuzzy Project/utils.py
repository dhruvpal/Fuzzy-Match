"""
Universal Matcher Utilities
===========================
Text normalization, data type detection, and multi-format export utilities.
"""

import re
import io
import pandas as pd
from typing import Any, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def normalize_generic_text(text: Any) -> str:
    """
    Standard domain-agnostic text normalizer:
    1. Converts to string and lowercase
    2. Strips leading and trailing whitespace
    3. Replaces punctuation with space (preserves alphanumeric characters)
    4. Collapses multiple spaces into a single space
    """
    if text is None or pd.isna(text):
        return ""
    
    s = str(text).lower().strip()
    if not s or s in ('nan', 'none', 'null', '<na>'):
        return ""
    
    # Replace non-alphanumeric punctuation with space
    s = re.sub(r'[^\w\s]', ' ', s)
    # Collapse multiple whitespace characters
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def detect_data_type(series: pd.Series) -> str:
    """
    Infers the general data type of a column ('numeric', 'date', 'text').
    """
    non_nulls = series.dropna().astype(str).str.strip()
    if len(non_nulls) == 0:
        return 'text'
    
    sample = non_nulls.head(50)
    
    # Check numeric
    numeric_count = 0
    for val in sample:
        clean_num = re.sub(r'[^\d.-]', '', val)
        if clean_num:
            try:
                float(clean_num)
                numeric_count += 1
            except ValueError:
                pass
    if numeric_count / len(sample) > 0.8:
        return 'numeric'
    
    # Check date
    date_count = 0
    for val in sample:
        try:
            pd.to_datetime(val, errors='raise', format='mixed')
            date_count += 1
        except Exception:
            pass
    if date_count / len(sample) > 0.8:
        return 'date'
    
    return 'text'


def export_dataframe_to_excel_bytes(
    df: pd.DataFrame,
    sheet_name: str = "Matching Results"
) -> bytes:
    """
    Exports a DataFrame to styled Excel (.xlsx) file bytes with auto-adjusted
    column widths, styled headers, and clean borders.
    """
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        worksheet = writer.sheets[sheet_name]

        # Style header
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        regular_font = Font(name="Segoe UI", size=9)
        thin_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )

        for col_idx, col in enumerate(df.columns, start=1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # Auto fit column width
            max_len = max(
                len(str(col)),
                max((len(str(v)) for v in df[col].dropna().head(100)), default=5)
            )
            col_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 45)

        # Style data rows
        for row in worksheet.iter_rows(min_row=2, max_row=len(df)+1, min_col=1, max_col=len(df.columns)):
            for cell in row:
                cell.font = regular_font
                cell.border = thin_border
                if isinstance(cell.value, (int, float)):
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

    output.seek(0)
    return output.getvalue()


def export_dataframe_to_csv_bytes(df: pd.DataFrame) -> bytes:
    """
    Exports a DataFrame to CSV bytes with UTF-8 BOM for Excel compatibility.
    """
    return df.to_csv(index=False, encoding='utf-8-sig').encode('utf-8-sig')
